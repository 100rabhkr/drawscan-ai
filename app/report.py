"""
Generate Layout Inspection Report Excel file matching URI's template.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=10)
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="center", horizontal="center")
LEFT_ALIGNMENT = Alignment(wrap_text=True, vertical="center", horizontal="left")


def _apply_border(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER


def generate_report(extraction: dict, output_path: str) -> str:
    """Generate a Layout Inspection Report Excel file from extraction data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Layout Inspection Report"

    # Column widths
    col_widths = {
        "A": 8,   # Sr No
        "B": 14,  # Type
        "C": 35,  # Description of Parameters
        "D": 22,  # Specified Value
        "E": 10,  # UOM
        "F": 14,  # Observed 1
        "G": 14,  # Observed 2
        "H": 14,  # Observed 3
        "I": 14,  # Observed 4
        "J": 14,  # Observed 5
        "K": 20,  # Inspection Method
        "L": 12,  # Remarks
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    tb = extraction.get("title_block", {})
    row = 1

    # --- Company Header ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1, value="UNITED RUBBER INDUSTRIES (I) PVT. LTD.")
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1, value="LAYOUT INSPECTION REPORT")
    cell.font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2

    # --- Title Block Info ---
    info_rows = [
        ("Customer:", tb.get("customer", "")),
        ("Part Name:", tb.get("part_name", "")),
        ("Drawing No:", tb.get("drawing_number", "")),
        ("Material:", tb.get("material", "")),
        ("Revision:", tb.get("revision", "")),
        ("Date:", tb.get("date", "")),
        ("General Tolerance:", tb.get("general_tolerance", "")),
    ]
    for label, value in info_rows:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=True, size=10)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        row += 1

    row += 1

    # --- Dimensions Table Header ---
    header_row = row
    headers = [
        "Sr\nNo",
        "Type",
        "Description of Parameters",
        "Specified Value\n(with Tolerance)",
        "UOM",
        "Observed\n1",
        "Observed\n2",
        "Observed\n3",
        "Observed\n4",
        "Observed\n5",
        "Inspection\nMethod",
        "Remarks",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGNMENT
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 35
    row += 1

    # --- Dimension Rows ---
    dimensions = extraction.get("dimensions", [])
    for dim in dimensions:
        dim_id = dim.get("id", "")
        dim_type = dim.get("type", "").capitalize()
        description = dim.get("description", "")
        nominal = dim.get("nominal_value", "")

        # Build specified value with tolerance
        tol_upper = dim.get("tolerance_upper") or ""
        tol_lower = dim.get("tolerance_lower") or ""
        if tol_upper and tol_lower and tol_upper == tol_lower:
            specified = f"{nominal} {tol_upper}"
        elif tol_upper and tol_lower:
            specified = f"{nominal} {tol_upper}/{tol_lower}"
        elif tol_upper:
            specified = f"{nominal} {tol_upper}"
        else:
            specified = str(nominal)

        unit = dim.get("unit", "mm")
        instrument = dim.get("suggested_instrument", "")

        ws.cell(row=row, column=1, value=dim_id).font = BODY_FONT
        ws.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
        ws.cell(row=row, column=2, value=dim_type).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = WRAP_ALIGNMENT
        ws.cell(row=row, column=3, value=description).font = BODY_FONT
        ws.cell(row=row, column=3).alignment = LEFT_ALIGNMENT
        ws.cell(row=row, column=4, value=specified).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row, column=4).alignment = WRAP_ALIGNMENT
        ws.cell(row=row, column=5, value=unit).font = BODY_FONT
        ws.cell(row=row, column=5).alignment = WRAP_ALIGNMENT
        # Observed value columns (F-J) left blank for manual entry
        for col in range(6, 11):
            ws.cell(row=row, column=col).font = BODY_FONT
            ws.cell(row=row, column=col).alignment = WRAP_ALIGNMENT
        ws.cell(row=row, column=11, value=instrument).font = BODY_FONT
        ws.cell(row=row, column=11).alignment = WRAP_ALIGNMENT
        ws.cell(row=row, column=12).font = BODY_FONT
        ws.cell(row=row, column=12).alignment = WRAP_ALIGNMENT

        _apply_border(ws, row, 1, 12)
        row += 1

    # --- GD&T Section ---
    gdt_items = extraction.get("gdt", [])
    if gdt_items:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        cell = ws.cell(row=row, column=1, value="GD&T (Geometric Dimensioning & Tolerancing)")
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = LEFT_ALIGNMENT
        _apply_border(ws, row, 1, 12)
        row += 1

        gdt_headers = [
            "Sr\nNo", "Symbol", "Applied To", "Tolerance Value",
            "Modifier", "Datum References", "", "", "", "", "", "Remarks"
        ]
        for col, header in enumerate(gdt_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = WRAP_ALIGNMENT
            cell.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=11)
        row += 1

        for gdt in gdt_items:
            ws.cell(row=row, column=1, value=gdt.get("id", "")).font = BODY_FONT
            ws.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
            ws.cell(row=row, column=2, value=gdt.get("symbol", "")).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = WRAP_ALIGNMENT
            ws.cell(row=row, column=3, value=gdt.get("applied_to", "")).font = BODY_FONT
            ws.cell(row=row, column=3).alignment = LEFT_ALIGNMENT
            ws.cell(row=row, column=4, value=gdt.get("tolerance_value", "")).font = BODY_FONT
            ws.cell(row=row, column=4).alignment = WRAP_ALIGNMENT
            ws.cell(row=row, column=5, value=gdt.get("modifier", "")).font = BODY_FONT
            ws.cell(row=row, column=5).alignment = WRAP_ALIGNMENT
            datums = ", ".join(gdt.get("datum_references", []))
            ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=11)
            ws.cell(row=row, column=6, value=datums).font = BODY_FONT
            ws.cell(row=row, column=6).alignment = WRAP_ALIGNMENT
            _apply_border(ws, row, 1, 12)
            row += 1

    # --- Notes Section ---
    notes = extraction.get("notes", [])
    if notes:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        cell = ws.cell(row=row, column=1, value="Notes")
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = LEFT_ALIGNMENT
        _apply_border(ws, row, 1, 12)
        row += 1

        for i, note in enumerate(notes, 1):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            ws.cell(row=row, column=1, value=f"{i}. {note}").font = BODY_FONT
            ws.cell(row=row, column=1).alignment = LEFT_ALIGNMENT
            _apply_border(ws, row, 1, 12)
            row += 1

    # --- Footer ---
    row += 2
    ws.cell(row=row, column=1, value="Prepared By:").font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=row, column=6, value="Verified By:").font = Font(name="Calibri", bold=True, size=10)

    # Print setup
    ws.print_area = f"A1:L{row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(output_path)
    return output_path
